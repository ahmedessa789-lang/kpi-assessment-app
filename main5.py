from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from openpyxl import Workbook
import sqlite3
from datetime import datetime

app = FastAPI(
    title="KPI Assessment API",
    version="1.0.0"
)

def init_db():
    conn = sqlite3.connect("kpi.db")
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS assessments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_name TEXT,
        overall_score REAL,
        overall_status TEXT,
        created_at TEXT
    )
    """)

    conn.commit()
    conn.close()

init_db()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:5501",
        "http://localhost:5501",
        "http://127.0.0.1:5500",
        "http://localhost:5500",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
class KPIInput(BaseModel):
    company_name: str = "Company"

    current_revenue: float
    previous_revenue: float
    leads_count: int
    converted_customers: int
    orders_count: int
    sales_target: float

    revenue: float
    cogs: float
    total_expenses: float
    accounts_receivable: float
    average_daily_sales: float
    cash_in: float
    cash_out: float

    total_orders: int
    on_time_orders: int
    completed_orders: int
    error_cases: int

    average_inventory: float
    accurate_count_items: int
    total_counted_items: int
    out_of_stock_cases: int
    total_sku_requests: int


@app.get("/")
def root():
    return {"message": "KPI Assessment API is running"}


def score_high_better(value, excellent, good, average):
    if value >= excellent:
        return 100, "green"
    elif value >= good:
        return 80, "yellow"
    elif value >= average:
        return 60, "yellow"
    return 40, "red"


def score_low_better(value, excellent, good, average):
    if value <= excellent:
        return 100, "green"
    elif value <= good:
        return 80, "yellow"
    elif value <= average:
        return 60, "yellow"
    return 40, "red"


def get_status(score):
    if score >= 85:
        return "Excellent", "green"
    elif score >= 70:
        return "Good", "yellow"
    return "Needs Improvement", "red"


@app.post("/kpi/calculate")
def calculate_kpi(data: KPIInput):
    # -------------------------
    # KPI Calculations
    # -------------------------
    sales_growth = ((data.current_revenue - data.previous_revenue) / data.previous_revenue * 100) if data.previous_revenue != 0 else 0
    conversion_rate = (data.converted_customers / data.leads_count * 100) if data.leads_count != 0 else 0
    average_order_value = (data.current_revenue / data.orders_count) if data.orders_count != 0 else 0
    target_achievement = (data.current_revenue / data.sales_target * 100) if data.sales_target != 0 else 0

    gross_profit_margin = ((data.revenue - data.cogs) / data.revenue * 100) if data.revenue != 0 else 0
    collection_period = (data.accounts_receivable / data.average_daily_sales) if data.average_daily_sales != 0 else 0
    expense_ratio = (data.total_expenses / data.revenue * 100) if data.revenue != 0 else 0
    cash_flow = data.cash_in - data.cash_out

    on_time_delivery = (data.on_time_orders / data.total_orders * 100) if data.total_orders != 0 else 0
    fulfillment_rate = (data.completed_orders / data.total_orders * 100) if data.total_orders != 0 else 0
    error_rate = (data.error_cases / data.total_orders * 100) if data.total_orders != 0 else 0

    inventory_turnover = (data.cogs / data.average_inventory) if data.average_inventory != 0 else 0
    stock_accuracy = (data.accurate_count_items / data.total_counted_items * 100) if data.total_counted_items != 0 else 0
    out_of_stock_rate = (data.out_of_stock_cases / data.total_sku_requests * 100) if data.total_sku_requests != 0 else 0

    # -------------------------
    # Scoring
    # -------------------------
    sg_score, _ = score_high_better(sales_growth, 20, 10, 0)
    cr_score, _ = score_high_better(conversion_rate, 30, 20, 10)
    ta_score, _ = score_high_better(target_achievement, 100, 80, 60)

    gpm_score, _ = score_high_better(gross_profit_margin, 30, 20, 10)
    cp_score, _ = score_low_better(collection_period, 30, 45, 60)
    er_score, _ = score_low_better(expense_ratio, 30, 45, 60)
    cf_score = 100 if cash_flow > 0 else 40

    otd_score, _ = score_high_better(on_time_delivery, 95, 85, 75)
    fr_score, _ = score_high_better(fulfillment_rate, 95, 85, 75)
    err_score, _ = score_low_better(error_rate, 2, 5, 10)

    it_score, _ = score_high_better(inventory_turnover, 8, 5, 3)
    sa_score, _ = score_high_better(stock_accuracy, 98, 95, 90)
    os_score, _ = score_low_better(out_of_stock_rate, 2, 5, 10)

    sales_score = round((sg_score + cr_score + ta_score) / 3, 2)
    finance_score = round((gpm_score + cp_score + er_score + cf_score) / 4, 2)
    operations_score = round((otd_score + fr_score + err_score) / 3, 2)
    inventory_score = round((it_score + sa_score + os_score) / 3, 2)

    overall_score = round((sales_score + finance_score + operations_score + inventory_score) / 4, 2)

    sales_status, sales_color = get_status(sales_score)
    finance_status, finance_color = get_status(finance_score)
    operations_status, operations_color = get_status(operations_score)
    inventory_status, inventory_color = get_status(inventory_score)
    overall_status, overall_color = get_status(overall_score)

    # -------------------------
    # Recommendations
    # -------------------------
    recommendations = []

    if sales_score < 70:
        recommendations.append("Improve lead conversion and track sales target achievement weekly.")
    if finance_score < 70:
        recommendations.append("Review collection cycle, expense ratio, and strengthen cash flow control.")
    if operations_score < 70:
        recommendations.append("Improve on-time delivery and reduce process errors.")
    if inventory_score < 70:
        recommendations.append("Improve stock accuracy and reduce out-of-stock incidents.")

    if not recommendations:
        recommendations.append("Overall performance is stable. Focus on continuous monitoring and incremental improvement.")

    return {
        "company_name": data.company_name,
        "sales": {
            "score": sales_score,
            "status": sales_status,
            "color": sales_color,
            "kpis": {
                "sales_growth": round(sales_growth, 2),
                "conversion_rate": round(conversion_rate, 2),
                "average_order_value": round(average_order_value, 2),
                "target_achievement": round(target_achievement, 2),
            },
        },
        "finance": {
            "score": finance_score,
            "status": finance_status,
            "color": finance_color,
            "kpis": {
                "gross_profit_margin": round(gross_profit_margin, 2),
                "collection_period": round(collection_period, 2),
                "expense_ratio": round(expense_ratio, 2),
                "cash_flow": round(cash_flow, 2),
            },
        },
        "operations": {
            "score": operations_score,
            "status": operations_status,
            "color": operations_color,
            "kpis": {
                "on_time_delivery": round(on_time_delivery, 2),
                "fulfillment_rate": round(fulfillment_rate, 2),
                "error_rate": round(error_rate, 2),
            },
        },
        "inventory": {
            "score": inventory_score,
            "status": inventory_status,
            "color": inventory_color,
            "kpis": {
                "inventory_turnover": round(inventory_turnover, 2),
                "stock_accuracy": round(stock_accuracy, 2),
                "out_of_stock_rate": round(out_of_stock_rate, 2),
            },
        },
        "overall": {
            "score": overall_score,
            "status": overall_status,
            "color": overall_color,
        },
        "recommendations": recommendations,
    }
# Fake Database 
assessments_db = []


@app.post("/kpi/save")
def save_assessment(data: KPIInput):
    result = calculate_kpi(data)

    conn = sqlite3.connect("kpi.db")
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO assessments
        (company_name, overall_score, overall_status, created_at)
        VALUES (?, ?, ?, ?)
    """, (
        data.company_name,
        result["overall"]["score"],
        result["overall"]["status"],
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ))

    conn.commit()
    conn.close()

    return {"message": "Assessment saved successfully"}


@app.get("/kpi/list")
def list_assessments():
    conn = sqlite3.connect("kpi.db")
    cursor = conn.cursor()

    cursor.execute("SELECT company_name FROM assessments")
    rows = cursor.fetchall()

    conn.close()

    return {"companies": [row[0] for row in rows]}


@app.get("/kpi/company/{name}")
def get_company(name: str):
    for item in assessments_db:
        if item["company_name"] == name:
            return item["result"]

    return {"error": "Company not found"}
@app.get("/kpi/export/{company_name}")
def export_excel(company_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Report"

    ws["A1"] = "Company Name"
    ws["B1"] = company_name

    ws["A3"] = "Generated By"
    ws["B3"] = "KPI Assessment Dashboard"

    ws["A5"] = "Status"
    ws["B5"] = "Exported Successfully"

    filename = f"{company_name}_KPI_Report.xlsx"
    wb.save(filename)

    return FileResponse(
        path=filename,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )